"""
excel_exporter.py
=================
Generates multi-sheet Excel files that mirror the exact structure of the
reference estimation files analysed from the Estimation folder.

AMC Outdoor  → sheets: Summary, Labour, Equipment, PPE, Consumables,
                        Subcontractor, Admin, Other, Assets
AMC Indoor   → sheets: CE - RF
Project Out  → sheets: CE - SK, BOQ
Project In   → sheets: CE - RF
"""

import io
import datetime
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from price_logic import MARGIN_PCT, OVERHEAD_PCT, NEGO_PCT

# ─── SHARED STYLE HELPERS ────────────────────────────────────────────────────

NAVY   = "1F3864"
TEAL   = "00B0A0"
SILVER = "D6DCE4"
WHITE  = "FFFFFF"
LIGHT  = "EBF3FB"
AMBER  = "FFC000"
GREEN  = "70AD47"
DARK   = "0A0F1E"

def _font(bold=False, size=11, color=WHITE, name="Calibri", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _border(style="thin"):
    s = Side(style=style, color="B0B8C1")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _money(value):
    """Format as AED currency string."""
    if value is None:
        return ""
    return f"AED {value:,.2f}"

def _pct(value):
    return f"{value*100:.0f}%"

def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def _cell(ws, row, col, value="", bold=False, size=11, fg=DARK, bg=None,
          align="left", wrap=False, italic=False, border=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", size=size, bold=bold, color=fg, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = _fill(bg)
    if border:
        c.border = _border()
    if number_format:
        c.number_format = number_format
    return c

def _header_row(ws, row, cols_values, bg=NAVY, fg=WHITE, size=11,
                bold=True, height=22):
    """Write a full header row with dark background."""
    for col, val in enumerate(cols_values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Calibri", size=size, bold=bold, color=fg)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[row].height = height

def _data_row(ws, row, cols_values, bg=None, bold=False, align_map=None,
              number_cols=None, height=18):
    """Write a data row with optional per-column alignment and number format."""
    align_map   = align_map   or {}
    number_cols = number_cols or {}
    for col, val in enumerate(cols_values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Calibri", size=10, bold=bold, color=DARK)
        c.alignment = Alignment(
            horizontal=align_map.get(col, "left"),
            vertical="center", wrap_text=True)
        c.border = _border()
        if bg:
            c.fill = _fill(bg)
        if col in number_cols:
            c.number_format = number_cols[col]
    ws.row_dimensions[row].height = height

def _section_title(ws, row, text, ncols=8, bg=SILVER):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Calibri", size=10, bold=True, color=NAVY)
    c.fill      = _fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = _border()
    ws.row_dimensions[row].height = 16

def _summary_title(ws, company, site, client, ref, date_str):
    """Write the standard 2-row company / pricing schedule title."""
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1,
                value="PRICING SCHEDULE  -  ITALIAN PLANTERS LLC  (Integrated Landscaping Solutions)")
    c.font      = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill      = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    c = ws.cell(row=2, column=1, value=f"SUMMARY  -  {site.upper()}")
    c.font      = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill      = _fill(TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20


# ─── SHEET BUILDERS ──────────────────────────────────────────────────────────

def _build_summary_sheet(wb, result, meta):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    _summary_title(ws, "Italian Planters LLC", meta["site"], meta["client"],
                   meta.get("ref", ""), meta.get("date", ""))

    # Info block
    info = [
        ("TENDER REFERENCE", "RFP", "PROJECT",
         "Operation and Maintenance of Site Landscape"),
        ("CLIENT NAME", meta["client"], "DATE (DD-MM-YY)",
         datetime.date.today().strftime("%d-%m-%Y")),
    ]
    for r, row in enumerate(info, 4):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = Font(name="Calibri", size=10,
                                  bold=(c % 2 == 1), color=DARK)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 16

    # Column headers (row 7)
    R = 7
    _header_row(ws, R, ["COST ELEMENT", "", "MOBILISATION",
                         "YEAR 1", "YEAR 2", "YEAR 3", "TOTAL", "%"])
    ws.row_dimensions[R].height = 22

    # Data rows
    NF = '#,##0.00'
    PCT = '0.00%'
    rows_data = [
        ("MOBILISATION",           "", "=0",
         "=0", "=0", "=0", "=SUM(C8:F8)", "=C8/D$17"),
        ("LABOUR DIRECT",          "", "=0",
         "=Labour!I19", "=D9", "=E9*1.05", "=SUM(C9:F9)", "=D9/D$17"),
        ("LABOUR INDIRECT",        "", "=0",
         "=Labour!I20", "=D10", "=E10*1.05", "=SUM(C10:F10)", "=D10/D$17"),
        ("CAPITAL EQUIPMENT",      "", "=0",
         "=Equipment!G13", "=D11", "=E11", "=SUM(C11:F11)", "=D11/D$17"),
        ("UNIFORM / PPE",          "", "=0",
         "=PPE!F14", "=D12", "=E12*1.05", "=SUM(C12:F12)", "=D12/D$17"),
        ("CONSUMABLES",            "", "=0",
         "=Consumables!H18", "=D13", "=E13", "=SUM(C13:F13)", "=D13/D$17"),
        ("SUBCONTRACTOR",          "", "=0",
         "=Subcontractor!G8", "=D14", "=E14*1.05", "=SUM(C14:F14)", "=D14/D$17"),
        ("ADMIN (DIRECT COST)",    "", "=0",
         "=0", "=0", "=0", "=SUM(C15:F15)", "=D15/D$17"),
        ("OTHER (INSURANCE)",      "", "=0",
         "=Other!H6", "=D16", "=E16*1.05", "=SUM(C16:F16)", "=D16/D$17"),
    ]

    num_cols = {3: NF, 4: NF, 5: NF, 6: NF, 7: NF, 8: PCT}
    for i, row_d in enumerate(rows_data):
        rr = R + 1 + i
        _data_row(ws, rr, row_d, align_map={3:"right",4:"right",5:"right",
                                             6:"right",7:"right",8:"right"},
                  number_cols=num_cols, height=17)

    # Subtotal
    ST = R + 1 + len(rows_data)
    _data_row(ws, ST, ["", "", "=SUM(C8:C16)",
                        "=SUM(D8:D16)", "=SUM(E8:E16)",
                        "=SUM(F8:F16)",
                        "=SUM(G8:G16)",
                        ""],
              bg=LIGHT, bold=True,
              align_map={3:"right",4:"right",5:"right",6:"right",7:"right"},
              number_cols={3:NF,4:NF,5:NF,6:NF,7:NF}, height=19)

    # Markup block
    margin_pct = result.get("margin_pct", MARGIN_PCT)
    overhead_pct = result.get("overhead_pct", OVERHEAD_PCT)
    nego_pct = result.get("nego_pct", NEGO_PCT)

    markup_rows = [
        ("Margin",      margin_pct, "=0", "=D$17*B18",   "=E$17*B18",   "=F$17*B18",   "=SUM(D18:F18)",   ""),
        ("Overhead",    overhead_pct, "=0", "=D$17*B19",  "=E$17*B19",   "=F$17*B19",   "=SUM(D19:F19)",   ""),
        ("Negotiation", nego_pct, "=0", "=(D$17-D$16)*B20", "=(E$17-E$16)*B20", "=(F$17-F$16)*B20", "=SUM(D20:F20)", ""),
    ]
    for i, row_d in enumerate(markup_rows):
        rr = ST + 1 + i
        _data_row(ws, rr, row_d,
                  align_map={2:"right",3:"right",4:"right",5:"right",6:"right",7:"right"},
                  number_cols={2:PCT,3:NF,4:NF,5:NF,6:NF,7:NF}, height=17)

    MR = ST + 4  # markup subtotal row
    _data_row(ws, MR, ["", "=SUM(B18:B20)", "=0",
                        "=SUM(D18:D20)", "=SUM(E18:E20)", "=SUM(F18:F20)",
                        "=SUM(G18:G20)", ""],
              bg=LIGHT, bold=True,
              align_map={2:"right",3:"right",4:"right",5:"right",6:"right",7:"right"},
              number_cols={2:PCT,3:NF,4:NF,5:NF,6:NF,7:NF}, height=19)

    # Grand totals
    GR = MR + 2
    _data_row(ws, GR, ["", "", "=0",
                        "=D$17+D$21", "=E$17+E$21",
                        "=F$17+F$21",
                        "=SUM(D23:F23)", ""],
              bg=NAVY, bold=True,
              align_map={3:"right",4:"right",5:"right",6:"right",7:"right"},
              number_cols={3:NF,4:NF,5:NF,6:NF,7:NF}, height=22)
    for col in range(1, 9):
        c = ws.cell(row=GR, column=col)
        c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    # Monthly row
    MN = GR + 1
    _data_row(ws, MN, ["MONTHLY RATE (incl. VAT 5%)", "", "",
                        "=(D23/12)*1.05", "=(E23/12)*1.05",
                        "=(F23/12)*1.05", "", ""],
              bg=TEAL, bold=True,
              align_map={4:"right",5:"right",6:"right"},
              number_cols={4:NF,5:NF,6:NF}, height=20)
    for col in range(1, 9):
        c = ws.cell(row=MN, column=col)
        c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    # Column widths
    widths = [30, 8, 14, 16, 16, 16, 18, 8]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)

    return ws


def _build_labour_sheet(wb, result, meta):
    ws = wb.create_sheet("Labour")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1, value="PRICING SCHEDULE")
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:I2")
    c = ws.cell(row=2, column=1, value="LABOUR DIRECT & INDIRECT")
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = _fill(TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    _header_row(ws, 3, ["Category", "Job Title", "No. FTE Staff",
                         "Basic Salary", "Fixed OT Rate", "Housing Allow.",
                         "Transport Allow.", "Food Allow.", "MONTHLY TOTAL"],
                height=32)

    r = result
    fte = r["fte_required"]
    # Approximate breakdown from analysed sheets
    direct_staff = [
        ("Direct Site Ops", "Gardeners - Labour",   round(fte*0.31, 4), 500,  156.25,  0, 0, 443.75),
        ("",                "Gardeners - Experts",  round(fte*0.62, 4), 675,  210.94,  0, 0, 464.91),
        ("",                "Irrigation Technician",round(fte*0.02, 4), 900,  281.25,  0, 0, 618.74),
        ("",                "Palm Tree Specialist",  round(fte*0.04, 4), 900, 281.25,  0, 0, 618.74),
        ("",                "Working Foreman",       round(fte*0.04, 4),2000, 1250.00, 0, 0, 750.00),
    ]
    indirect_staff = [
        ("Ops & Mgmt", "Site Superintendent", round(fte*0.01, 4), 1875, 0, 0,    0,    1875),
        ("",           "Admin",               round(fte*0.01, 4), 2000, 0, 800, 300,    900),
        ("",           "Landscaping Engineer",round(fte*0.01, 4), 4000, 0, 2000,  0,   2000),
        ("",           "HSE Engineer",        round(fte*0.005,4), 4000, 0, 2000,  0,   2000),
    ]

    NF = '#,##0.00'
    row = 4
    _section_title(ws, row, "LABOUR DIRECT", ncols=9)
    row += 1
    for staff in direct_staff:
        _data_row(ws, row, staff,
                  align_map={3:"right",4:"right",5:"right",6:"right",7:"right",8:"right",9:"right"},
                  number_cols={3:'0.0000',4:NF,5:NF,6:NF,7:NF,8:NF,9:NF}, height=17)
        row += 1

    # Direct subtotal
    _data_row(ws, row, ["", "TOTAL DIRECT", round(fte*1.03,4),
                         "", "", "", "", "",
                         round(r["labour_direct_y1"]/12, 2)],
              bg=LIGHT, bold=True,
              align_map={3:"right",9:"right"},
              number_cols={3:'0.0000',9:NF}, height=19)
    row += 2

    _section_title(ws, row, "LABOUR INDIRECT", ncols=9)
    row += 1
    for staff in indirect_staff:
        _data_row(ws, row, staff,
                  align_map={3:"right",4:"right",5:"right",6:"right",7:"right",8:"right",9:"right"},
                  number_cols={3:'0.0000',4:NF,5:NF,6:NF,7:NF,8:NF,9:NF}, height=17)
        row += 1

    _data_row(ws, row, ["", "TOTAL INDIRECT", "", "", "", "", "", "",
                         round(r["labour_indirect_y1"]/12, 2)],
              bg=LIGHT, bold=True,
              align_map={9:"right"}, number_cols={9:NF}, height=19)
    row += 2

    # Annual summary
    _data_row(ws, row,
              ["ANNUAL LABOUR DIRECT",  "", "", "", "", "", "", "",
               r["labour_direct_y1"]],
              bg=NAVY, bold=True,
              align_map={9:"right"}, number_cols={9:NF}, height=20)
    for col in range(1,10):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    row += 1
    _data_row(ws, row,
              ["ANNUAL LABOUR INDIRECT",  "", "", "", "", "", "", "",
               r["labour_indirect_y1"]],
              bg=TEAL, bold=True,
              align_map={9:"right"}, number_cols={9:NF}, height=20)
    for col in range(1,10):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    widths = [20,28,12,12,12,12,12,12,14]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    return ws


def _build_equipment_sheet(wb, result, meta):
    ws = wb.create_sheet("Equipment")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value="PRICING SCHEDULE — CAPITAL EQUIPMENT")
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    _header_row(ws, 2, ["DETAILS", "BRAND", "TYPE",
                         "OWNED/LEASED", "QTY", "UNIT COST (AED)",
                         "TOTAL COST (AED)", "NOTES"], height=24)

    NF = '#,##0.00'
    equipment = [
        ("Hand Held Blower Machine",    "Maruyama",    "OPEX",  "OWNED", 0.2,  1800,  360),
        ("Hedge Trimmer",               "Stihl HS-82R","OPEX",  "OWNED", 0.2,  2600,  520),
        ("String Trimmer Machine",      "EFCO Stark",  "OPEX",  "OWNED", 0.2,  1732,  346.40),
        ("Backpack Sprayer",            "",            "CAPEX", "OWNED", 0.1,  1800,  180),
        ("Garden Tools Box",            "",            "OPEX",  "OWNED", 0.5,  300,   150),
        ("Irrigation Tech Tools Box",   "",            "OPEX",  "OWNED", 0.05, 1000,  50),
        ("Wheel Barrow",                "",            "OPEX",  "OWNED", 0.25, 150,   37.5),
        ("Hose",                        "Maimoosa",    "OPEX",  "OWNED", 1,    300,   300),
        ("Machines Consumables",        "",            "OPEX",  "OWNED", 1,    43.33, 43.33),
        ("Machines Maintenance",        "",            "OPEX",  "OWNED", 1,    21.38, 21.38),
    ]
    row = 3
    for eq in equipment:
        _data_row(ws, row, eq,
                  align_map={5:"right",6:"right",7:"right"},
                  number_cols={5:'0.00',6:NF,7:NF}, height=17)
        row += 1

    # Total
    total_eq = result["equipment_y1"]
    _data_row(ws, row, ["TOTAL EQUIPMENT (ANNUAL)", "", "", "", "", "",
                         total_eq, ""],
              bg=NAVY, bold=True,
              align_map={7:"right"}, number_cols={7:NF}, height=20)
    for col in range(1,9):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    widths = [28, 20, 10, 12, 8, 18, 18, 20]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    return ws


def _build_ppe_sheet(wb, result, meta):
    ws = wb.create_sheet("PPE")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value="PRICING SCHEDULE — UNIFORM / PPE")
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    _header_row(ws, 2, ["DETAILS", "TYPE", "UOM", "QTY",
                         "UNIT COST (AED)", "TOTAL COST (AED)", "NOTES"], height=22)
    NF = '#,##0.00'
    items = [
        ("Goggles",              "PPE",   "pair", 0.5, 10,  5),
        ("Helmet",               "PPE",   "nos",  0.5, 20,  10),
        ("Safety Shoes",         "PPE",   "pair", 0.5, 60,  30),
        ("Gumboots",             "PPE",   "pair", 0.5, 20,  10),
        ("Safety Jacket",        "PPE",   "nos",  0.5, 5,   2.5),
        ("Hand Gloves - Soft",   "PPE",   "pair", 0.5, 5,   2.5),
        ("Corrosive Gloves",     "PPE",   "pair", 0.5, 5,   2.5),
        ("Hand Gloves - Hard",   "PPE",   "pair", 0.5, 6,   3),
        ("Respiratory Mask",     "PPE",   "nos",  0.5, 40,  20),
        ("Warning Tape",         "Safety","roll",  2,  25,  50),
        ("Safety Cone",          "Safety","nos",   2,  30,  60),
    ]
    row = 3
    for it in items:
        _data_row(ws, row, it,
                  align_map={4:"right",5:"right",6:"right"},
                  number_cols={4:'0.00',5:NF,6:NF}, height=17)
        row += 1

    _data_row(ws, row, ["TOTAL PPE (ANNUAL)", "", "", "", "",
                         result["ppe_y1"], ""],
              bg=NAVY, bold=True, align_map={6:"right"}, number_cols={6:NF}, height=20)
    for col in range(1,8):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    widths = [24, 10, 8, 8, 16, 16, 20]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    return ws


def _build_consumables_sheet(wb, result, meta):
    ws = wb.create_sheet("Consumables")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value="PRICING SCHEDULE — CONSUMABLES")
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    _header_row(ws, 2, ["DESCRIPTION", "UOM", "UNIT COST",
                         "QTY / MONTH", "NO. MONTHS / YR",
                         "YEAR 1 (AED)", "MARGIN+OH (AED)", "TOTAL (AED)"], height=28)
    NF = '#,##0.00'
    r = result
    cons = [
        ("In-organic Fertilizer (NPK)",       "kg",      175,   0.17,  2,  60.06),
        ("In-organic Fertilizer Urea/Ammonium","kg",      100,   0.17,  2,  34.32),
        ("Iron Chelate",                       "kg",      50,    0.17,  1,   8.58),
        ("Organic Fertilizer (Compost)",       "kg",      8,     5.0,   1,  40.00),
        ("Insecticide / Pesticide",            "ml",      125,   0.1,   2,  25.00),
        ("Fungicide / Pesticide",              "ml",      110,   0.1,   2,  22.00),
        ("Herbicide (Garlon)",                 "ml",      200,   0.1,   1,  20.00),
        ("Other Pesticide",                    "ml",      130,   0.04,  1,   5.57),
        ("Garbage Bags",                       "roll",    80,    0.25, 12, 240.00),
        ("Irrigation Rectification Allowance", "lumpsum", 50,    1,     6, 300.00),
        ("Seasonal Flowers Replacement",       "nos",     100,   1.0,   3, 300.00),
        ("Outdoor Plants Mortality Replacement","lumpsum", 50,   1,     6, 300.00),
    ]
    row = 3
    sub = 0
    for it in cons:
        _data_row(ws, row, (it[0],it[1],it[2],it[3],it[4],it[5],"",""),
                  align_map={3:"right",4:"right",5:"right",6:"right"},
                  number_cols={3:NF,4:'0.00',6:NF}, height=17)
        sub += it[5]
        row += 1

    # markup on consumables
    margin   = sub * 0.10
    overhead = sub * 0.10
    total_c  = r["consumables_y1"]

    _data_row(ws, row, ["", "", "", "", "Sub-Total",  sub, "", ""],
              bg=LIGHT, bold=True, align_map={5:"right",6:"right"},
              number_cols={6:NF}, height=19)
    row += 1
    _data_row(ws, row, ["", "", "", "", "Margin 10%", margin, "", ""],
              align_map={5:"right",6:"right"}, number_cols={6:NF}, height=17)
    row += 1
    _data_row(ws, row, ["", "", "", "", "Overhead 10%", overhead, "", ""],
              align_map={5:"right",6:"right"}, number_cols={6:NF}, height=17)
    row += 1
    _data_row(ws, row, ["TOTAL CONSUMABLES (ANNUAL)", "", "", "", "",
                         total_c, "", total_c],
              bg=NAVY, bold=True,
              align_map={6:"right",8:"right"}, number_cols={6:NF,8:NF}, height=20)
    for col in range(1,9):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    widths = [36, 10, 12, 12, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    return ws


def _build_subcontractor_sheet(wb, result, meta):
    ws = wb.create_sheet("Subcontractor")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value="PRICING SCHEDULE — SUBCONTRACTOR")
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    _header_row(ws, 2, ["SERVICE DESCRIPTION", "CATEGORY", "S/C NAME",
                         "UOM", "UNIT COST", "QTY / YEAR",
                         "YEAR 1 (AED)", "YEAR 2 (AED)"], height=22)
    NF = '#,##0.00'
    sub = result["subcontractor_y1"]
    items = [
        ("Hard Pruning Waste Removal", "All", "", "lumpsum", sub*0.8, 1, sub*0.8, sub*0.8),
        ("Soil pH & Fertility Test",   "All", "", "lumpsum", sub*0.2, 1, sub*0.2, sub*0.2),
    ]
    row = 3
    for it in items:
        _data_row(ws, row, it,
                  align_map={5:"right",6:"right",7:"right",8:"right"},
                  number_cols={5:NF,7:NF,8:NF}, height=17)
        row += 1

    _data_row(ws, row, ["", "", "", "", "", "", sub, sub],
              bg=LIGHT, bold=True, align_map={7:"right",8:"right"},
              number_cols={7:NF,8:NF}, height=19)
    row += 1
    _data_row(ws, row, ["", "", "", "Margin",   "0.10", "", sub*0.1, sub*0.1],
              align_map={5:"right",7:"right",8:"right"},
              number_cols={7:NF,8:NF}, height=17)
    row += 1
    _data_row(ws, row, ["", "", "", "Overhead",  "0.10", "", sub*0.1, sub*0.1],
              align_map={5:"right",7:"right",8:"right"},
              number_cols={7:NF,8:NF}, height=17)
    row += 1
    total_s = sub * 1.2
    _data_row(ws, row, ["TOTAL SUBCONTRACTOR (ANNUAL)", "", "", "", "", "",
                         total_s, total_s],
              bg=NAVY, bold=True,
              align_map={7:"right",8:"right"}, number_cols={7:NF,8:NF}, height=20)
    for col in range(1,9):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    widths = [32, 12, 14, 12, 12, 12, 16, 16]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    return ws


def _build_other_sheet(wb, result, meta):
    ws = wb.create_sheet("Other")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value="PRICING SCHEDULE — OTHER (INSURANCE & MISC)")
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    _header_row(ws, 2, ["DETAILS", "ITEM DESCRIPTION", "UOM",
                         "QTY", "UNIT COST", "TOTAL COST",
                         "FREQUENCY/YR", "YEAR 1 (AED)"], height=22)
    NF = '#,##0.00'
    items = [
        ("Third Party Insurance", "", "",      0.010, 7000, 70,    1,  70),
        ("Workmen Compensation",  "", "",      0.020, 2500, 50,    1,  50),
        ("PI Cover",              "", "",      0.005,11500, 57.5,  1, 57.5),
    ]
    row = 3
    for it in items:
        _data_row(ws, row, it,
                  align_map={4:"right",5:"right",6:"right",8:"right"},
                  number_cols={4:'0.000',5:NF,6:NF,8:NF}, height=17)
        row += 1

    oth = result["other_y1"]
    _data_row(ws, row, ["", "", "", "", "", "", "TOTAL", oth],
              bg=NAVY, bold=True, align_map={7:"right",8:"right"},
              number_cols={8:NF}, height=20)
    for col in range(1,9):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    widths = [28, 20, 8, 10, 12, 12, 14, 14]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    return ws


def _build_assets_sheet(wb, result, meta, config):
    ws = wb.create_sheet("Assets")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1,
                value=f"SITE ASSET REGISTER — {meta['site'].upper()}")
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    info = [
        ("Total Landscape Area (Equiv.)", f"{result['equiv_area_m2']} m²"),
        ("Calculated FTE Required",       str(result['fte_required'])),
        ("Site / Location",               meta.get('site','')),
        ("Client",                        meta.get('client','')),
    ]
    for r, (k, v) in enumerate(info, 2):
        ws.cell(row=r, column=1, value=k).font = Font(name="Calibri", size=10, bold=True, color=DARK)
        ws.cell(row=r, column=2, value=v).font = Font(name="Calibri", size=10, color=DARK)
        ws.row_dimensions[r].height = 16

    _header_row(ws, 7, ["S/N", "DESCRIPTION", "PALMS (nos)",
                         "TREES (nos)", "SHRUBS (m²)", "GC (m²)",
                         "LAWN (m²)", "SEASONAL FLOWERS (nos)"], height=28)

    _data_row(ws, 8, [1, meta.get('site','Site'),
                       config.get('palms',0), config.get('trees',0),
                       config.get('shrubs_m2',0), config.get('gc_m2',0),
                       config.get('lawn_m2',0), config.get('seasonal_flowers',0)],
              align_map={3:"right",4:"right",5:"right",6:"right",7:"right",8:"right"},
              height=18)
    _data_row(ws, 9, ["Total", "",
                       config.get('palms',0), config.get('trees',0),
                       config.get('shrubs_m2',0), config.get('gc_m2',0),
                       config.get('lawn_m2',0), config.get('seasonal_flowers',0)],
              bg=LIGHT, bold=True,
              align_map={3:"right",4:"right",5:"right",6:"right",7:"right",8:"right"},
              height=18)

    # Equiv area breakdown
    row = 11
    _section_title(ws, row, "EQUIVALENT AREA BREAKDOWN", ncols=8)
    _data_row(ws, 12, ["", "Equivalent Area (m²)",
                        config.get('palms',0)*13.5, config.get('trees',0)*6.0,
                        config.get('shrubs_m2',0), config.get('gc_m2',0),
                        config.get('lawn_m2',0), ""],
              align_map={3:"right",4:"right",5:"right",6:"right",7:"right"},
              number_cols={3:'#,##0.0',4:'#,##0.0',5:'#,##0.0',6:'#,##0.0',7:'#,##0.0'},
              height=17)
    _data_row(ws, 13, ["", "Manpower Requirement (FTE)",
                        config.get('palms',0)*13.5*0.00025,
                        config.get('trees',0)*6.0*0.00025,
                        config.get('shrubs_m2',0)*0.00025,
                        config.get('gc_m2',0)*0.00025,
                        config.get('lawn_m2',0)*0.00025, ""],
              align_map={3:"right",4:"right",5:"right",6:"right",7:"right"},
              number_cols={3:'0.0000',4:'0.0000',5:'0.0000',6:'0.0000',7:'0.0000'},
              height=17)

    widths = [6, 30, 14, 14, 14, 14, 14, 20]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    return ws


def _build_ce_indoor_amc(wb, result, meta):
    ws = wb.create_sheet("CE - RF")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1,
                value="Italian Planters LLC  ( Integrated Landscaping Solutions )")
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:G2")
    c = ws.cell(row=2, column=1, value="COST ESTIMATION MODEL")
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = _fill(TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # File / date info
    ws.cell(row=3, column=1, value=f"CE - {meta['client']} - {meta['site']}").font = Font(name="Calibri",size=10,bold=True,color=DARK)
    ws.cell(row=4, column=1, value=f"As on {datetime.date.today().strftime('%B %d, %Y')}").font = Font(name="Calibri",size=10,color=DARK)

    _header_row(ws, 5, ["", "SL", "DESCRIPTION", "QTY (Nos)",
                         "Unit Rate (AED)", "MONTHLY CHARGES", "ANNUAL AMOUNT"],
                height=22)

    NF = '#,##0.00'
    row = 6
    for i, line in enumerate(result["line_details"], 1):
        _data_row(ws, row, ["", i, line["description"],
                             line["qty"], line["unit_rate"],
                             line["monthly_charge"], line["annual_charge"]],
                  align_map={4:"right",5:"right",6:"right",7:"right"},
                  number_cols={5:NF,6:NF,7:NF}, height=18)
        row += 1

    # Totals
    _data_row(ws, row, ["", "", "TOTAL AMOUNT", "",  "",
                         result["base_monthly"], result["base_annual"]],
              bg=NAVY, bold=True,
              align_map={6:"right",7:"right"}, number_cols={6:NF,7:NF}, height=20)
    for col in range(1,8):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    row += 1

    _data_row(ws, row, ["", "", "Discount Applied", "",
                         "", result["discount_monthly"], result["discount_monthly"]*12],
              align_map={6:"right",7:"right"}, number_cols={6:NF,7:NF}, height=17)
    row += 1
    _data_row(ws, row, ["", "", "Net Amount (after discount)", "",  "",
                         result["net_monthly"], result["net_annual"]],
              bg=LIGHT, bold=True,
              align_map={6:"right",7:"right"}, number_cols={6:NF,7:NF}, height=18)
    row += 1
    _data_row(ws, row, ["", "", "VAT 5%", "", "",
                         result["vat_monthly"], result["vat_monthly"]*12],
              align_map={6:"right",7:"right"}, number_cols={6:NF,7:NF}, height=17)
    row += 1
    _data_row(ws, row, ["", "", "Total Amount (incl. VAT)", "", "",
                         result["gross_monthly_incl_vat"],
                         result["gross_monthly_incl_vat"]*12],
              bg=TEAL, bold=True,
              align_map={6:"right",7:"right"}, number_cols={6:NF,7:NF}, height=20)
    for col in range(1,8):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    # T&C note
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1,
                value="Note: Mortality replacement of Indoor plants is included in above maintenance charges.")
    c.font = Font(name="Calibri", size=9, italic=True, color="666666")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    widths = [4, 6, 40, 10, 16, 18, 18]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    return ws


def _build_ce_project_outdoor(wb, result, meta):
    ws = wb.create_sheet("CE - SK")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:P1")
    c = ws.cell(row=1, column=1,
                value="Italian Planters LLC  ( Integrated Landscaping Solutions )")
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:P2")
    c = ws.cell(row=2, column=1, value="COST ESTIMATION MODEL")
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = _fill(TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.cell(row=3, column=1, value=f"CE - {meta['client']} - {meta['site']}").font = Font(name="Calibri",size=10,bold=True,color=DARK)
    ws.cell(row=4, column=1, value=f"As on {datetime.date.today().strftime('%B %d, %Y')}").font = Font(name="Calibri",size=10,color=DARK)

    # Read markup percentages
    oh_material_pct = result.get("oh_material_pct", 0.20)
    oh_labor_pct = result.get("oh_labor_pct", 0.20)
    profit_material_pct = result.get("profit_material_pct", 0.20)
    profit_labor_pct = result.get("profit_labor_pct", 0.20)
    nego_pct = result.get("nego_pct", 0.10)

    oh_mat_label = f"OH ({oh_material_pct*100:.0f}%)\n(Material+Plant)"
    oh_lab_label = f"OH ({oh_labor_pct*100:.0f}%)\n(Labour)"
    prof_mat_label = f"PROFIT ({profit_material_pct*100:.0f}%)\n(Material+Plant)"
    prof_lab_label = f"PROFIT ({profit_labor_pct*100:.0f}%)\n(Labour)"
    nego_label = f"NEGO.    ({nego_pct*100:.0f}%)"

    _header_row(ws, 5, [
        "S/N", "DESCRIPTION OF WORKS", "QTY.", "UNIT",
        "COST", "LABOR", "SUPRV.", "PLANT",
        oh_mat_label, oh_lab_label, prof_mat_label, prof_lab_label,
        nego_label, "TOTAL", "RATE", "AMOUNT"
    ], height=32)

    NF = '#,##0.00'
    row = 6
    for i, item in enumerate(result["manifest"], 1):
        cols_values = [
            i,
            item["description"],
            item["qty"],
            item["unit"],
            item["cost"],
            item["labor"],
            item["suprv"],
            item["plant"],
            f"=(E{row}+H{row})*{oh_material_pct}",
            f"=(F{row}+G{row})*{oh_labor_pct}",
            f"=(E{row}+H{row})*{profit_material_pct}",
            f"=(F{row}+G{row})*{profit_labor_pct}",
            f"=(E{row}+F{row}+G{row}+H{row})*{nego_pct}",
            f"=SUM(E{row}:M{row})",
            f"=CEILING(N{row},1)",
            f"=C{row}*O{row}"
        ]
        _data_row(ws, row, cols_values,
                  align_map={3:"right", 4:"center", 5:"right", 6:"right", 7:"right", 8:"right",
                             9:"right", 10:"right", 11:"right", 12:"right", 13:"right",
                             14:"right", 15:"right", 16:"right"},
                  number_cols={5:NF, 6:NF, 7:NF, 8:NF, 9:NF, 10:NF, 11:NF, 12:NF, 13:NF, 14:NF, 15:NF, 16:NF},
                  height=18)
        row += 1

    # Total block
    total_row = row
    _data_row(ws, total_row, ["", "Total Amount", "", "", "", "", "", "", "", "", "", "", "", "", "", f"=SUM(P6:P{total_row-1})"],
              bg=NAVY, bold=True, align_map={16:"right"}, number_cols={16:NF}, height=20)
    for col in range(1, 17):
        ws.cell(row=total_row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    disc_row = total_row + 1
    _data_row(ws, disc_row, ["", "Discount / Negotiation", "", "", "", "", "", "", "", "", "", "", "", "", "", result["discount"]],
              align_map={16:"right"}, number_cols={16:NF}, height=17)

    tax_row = total_row + 2
    _data_row(ws, tax_row, ["", "Taxable Base", "", "", "", "", "", "", "", "", "", "", "", "", "", f"=P{total_row}-P{disc_row}"],
              bg=LIGHT, bold=True, align_map={16:"right"}, number_cols={16:NF}, height=18)

    vat_row = total_row + 3
    _data_row(ws, vat_row, ["", "VAT 5%", "", "", "", "", "", "", "", "", "", "", "", "", "", f"=P{tax_row}*0.05"],
              align_map={16:"right"}, number_cols={16:NF}, height=17)

    final_row = total_row + 4
    _data_row(ws, final_row, ["", "FINAL QUOTED AMOUNT", "", "", "", "", "", "", "", "", "", "", "", "", "", f"=P{tax_row}+P{vat_row}"],
              bg=TEAL, bold=True, align_map={16:"right"}, number_cols={16:NF}, height=22)
    for col in range(1, 17):
        ws.cell(row=final_row, column=col).font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    row = final_row + 2

    # T&C section
    _section_title(ws, row, "TERMS & CONDITIONS", ncols=16, bg="FFF2CC")
    row += 1
    tc_list = [
        "Any other works which is not mentioned in the scope above are excluded.",
        "Any Work Permit / Statutory Authority charges are excluded.",
        "Any Civil works is excluded.",
        "Any connection of irrigation lines is excluded unless specified.",
        "Standard UAE VAT at 5% is applicable on all services.",
    ]
    for tc in tc_list:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
        c = ws.cell(row=row, column=1, value=f"• {tc}")
        c.font = Font(name="Calibri", size=9, italic=True, color="444444")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 15
        row += 1

    widths = [5, 38, 7, 8, 11, 11, 11, 11, 14, 14, 14, 14, 12, 12, 12, 15]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    return ws


def _build_ce_project_indoor(wb, result, meta):
    ws = wb.create_sheet("CE - RF")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:Q1")
    c = ws.cell(row=1, column=1,
                value="Italian Planters LLC  ( Integrated Landscaping Solutions )")
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:Q2")
    c = ws.cell(row=2, column=1, value="COST ESTIMATION MODEL")
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = _fill(TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.cell(row=3, column=1, value=f"CE - {meta['client']} - {meta['site']}").font = Font(name="Calibri",size=10,bold=True,color=DARK)
    ws.cell(row=4, column=1, value=f"As on {datetime.date.today().strftime('%B %d, %Y')}").font = Font(name="Calibri",size=10,color=DARK)

    # Read markup percentages
    oh_pot_pct = result.get("oh_pot_pct", 0.20)
    oh_plant_pct = result.get("oh_plant_pct", 0.20)
    profit_pct = result.get("profit_pct", 0.20)
    nego_pct = result.get("nego_pct", 0.10)

    _header_row(ws, 5, [
        "SL", "DESCRIPTION", "POT (AED)", "MAIN PLANT (AED)",
        "UNDER PLANTS (AED)", "SOIL & HYDROSTONE (AED)",
        "SUB-TOTAL POT", "SUB-TOTAL Plant+Soil+Mulch",
        f"Overhead ({oh_pot_pct*100:.0f}%) POT",
        f"Overhead ({oh_plant_pct*100:.0f}%) Plants/Soil",
        f"Profit ({profit_pct*100:.0f}%)",
        f"Negotiation ({nego_pct*100:.0f}%)",
        "Labor", "Delivery Charge", "TOTAL (AED)", "QTY", "AMOUNT (AED)"
    ], height=36)

    NF = '#,##0.00'
    row = 6
    for i, item in enumerate(result["manifest"], 1):
        cols_values = [
            i,
            item["description"],
            item["pot"],
            item["main_plant"],
            item["under_plants"],
            item["soil_hydrostone"],
            f"=C{row}",
            f"=D{row}+E{row}+F{row}",
            f"=G{row}*{oh_pot_pct}",
            f"=H{row}*{oh_plant_pct}",
            f"=(G{row}+H{row})*{profit_pct}",
            f"=(G{row}+H{row})*{nego_pct}",
            item["labor"],
            item["delivery"],
            f"=SUM(G{row}:N{row})",
            item["qty"],
            f"=P{row}*O{row}"
        ]
        _data_row(ws, row, cols_values,
                  align_map={3:"right", 4:"right", 5:"right", 6:"right",
                             7:"right", 8:"right", 9:"right", 10:"right",
                             11:"right", 12:"right", 13:"right", 14:"right",
                             15:"right", 16:"right", 17:"right"},
                  number_cols={3:NF, 4:NF, 5:NF, 6:NF, 7:NF, 8:NF, 9:NF, 10:NF,
                               11:NF, 12:NF, 13:NF, 14:NF, 15:NF, 17:NF},
                  height=18)
        row += 1

    total_row = row
    _data_row(ws, total_row, ["", "Total Amount", "", "", "", "", "", "", "", "", "", "", "", "", "", "", f"=SUM(Q6:Q{total_row-1})"],
              bg=NAVY, bold=True, align_map={17:"right"}, number_cols={17:NF}, height=20)
    for col in range(1, 18):
        ws.cell(row=total_row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    disc_row = total_row + 1
    _data_row(ws, disc_row, ["", "Discount", "", "", "", "", "", "", "", "", "", "", "", "", "", "", result["discount"]],
              align_map={17:"right"}, number_cols={17:NF}, height=17)

    tax_row = total_row + 2
    _data_row(ws, tax_row, ["", "Taxable Base", "", "", "", "", "", "", "", "", "", "", "", "", "", "", f"=Q{total_row}-Q{disc_row}"],
              bg=LIGHT, bold=True, align_map={17:"right"}, number_cols={17:NF}, height=18)

    vat_row = total_row + 3
    _data_row(ws, vat_row, ["", "VAT 5%", "", "", "", "", "", "", "", "", "", "", "", "", "", "", f"=Q{tax_row}*0.05"],
              align_map={17:"right"}, number_cols={17:NF}, height=17)

    final_row = total_row + 4
    _data_row(ws, final_row, ["", "FINAL QUOTED AMOUNT", "", "", "", "", "", "", "", "", "", "", "", "", "", "", f"=Q{tax_row}+Q{vat_row}"],
              bg=TEAL, bold=True, align_map={17:"right"}, number_cols={17:NF}, height=22)
    for col in range(1, 18):
        ws.cell(row=final_row, column=col).font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    row = final_row + 2

    # Markup reference block
    _section_title(ws, row, "MARKUP STRUCTURE REFERENCE", ncols=17, bg="FFF2CC")
    row += 1
    markup_info = [
        f"Overhead on Pot:        {oh_pot_pct*100:.0f}%  |  Overhead on Plant+Soil:  {oh_plant_pct*100:.0f}%",
        f"Profit Margin:          {profit_pct*100:.0f}%  |  Negotiation Buffer:       {nego_pct*100:.0f}%",
        "UAE VAT:                 5%  applied on taxable base",
    ]
    for line in markup_info:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
        c = ws.cell(row=row, column=1, value=line)
        c.font = Font(name="Calibri", size=9, italic=True, color="444444")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 15
        row += 1

    widths = [5, 36, 12, 14, 14, 18, 14, 22, 16, 16, 14, 14, 12, 12, 14, 8, 15]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    return ws


def _build_boq_sheet(wb, result, meta):
    """Second sheet for project files — BOQ / quantity breakdown."""
    ws = wb.create_sheet("BOQ")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value=f"BILL OF QUANTITIES — {meta['site'].upper()}")
    c.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    _header_row(ws, 2, ["S/N", "DESCRIPTION", "QTY", "UNIT",
                         "UNIT RATE (AED)", "AMOUNT (AED)", "REMARKS"], height=22)
    NF = '#,##0.00'
    row = 3
    items = result.get("manifest", [])
    for i, item in enumerate(items, 1):
        _data_row(ws, row,
                  [i, item.get("description",""),
                   item.get("qty",""), item.get("unit","nos"),
                   item.get("unit_rate",""), item.get("extended",""), ""],
                  align_map={3:"right",5:"right",6:"right"},
                  number_cols={5:NF,6:NF}, height=18)
        row += 1

    _data_row(ws, row, ["", "GRAND TOTAL", "", "", "",
                         result["gross_total"], ""],
              bg=NAVY, bold=True, align_map={6:"right"}, number_cols={6:NF}, height=20)
    for col in range(1,8):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True, color=WHITE)

    widths = [5, 40, 8, 8, 16, 16, 20]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    return ws


# ─── MAIN ENTRY POINT ────────────────────────────────────────────────────────

def generate_excel(job_type: str, result: dict, config: dict,
                   meta: dict) -> bytes:
    """
    Build a multi-sheet Excel workbook matching the reference file formats.
    Returns the workbook as bytes (for HTTP response or file save).

    Parameters
    ----------
    job_type : one of 'amc_outdoor', 'amc_indoor',
                       'project_outdoor', 'project_indoor'
    result   : output dict from price_logic.get_estimate()
    config   : the original config dict passed to get_estimate()
    meta     : {'site': str, 'client': str}
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    if job_type == "amc_outdoor":
        _build_summary_sheet(wb, result, meta)
        _build_labour_sheet(wb, result, meta)
        _build_equipment_sheet(wb, result, meta)
        _build_ppe_sheet(wb, result, meta)
        _build_consumables_sheet(wb, result, meta)
        _build_subcontractor_sheet(wb, result, meta)
        # Admin sheet (minimal)
        ws_admin = wb.create_sheet("Admin")
        ws_admin.merge_cells("A1:E1")
        c = ws_admin.cell(row=1, column=1, value="ADMIN (DIRECT COST — NOT INCLUDED IN MARKUP)")
        c.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_admin.cell(row=2, column=1, value="No admin costs applicable for this contract.").font = \
            Font(name="Calibri", size=10, italic=True, color="666666")
        ws_admin.row_dimensions[1].height = 24
        _build_other_sheet(wb, result, meta)
        _build_assets_sheet(wb, result, meta, config)

    elif job_type == "amc_indoor":
        _build_ce_indoor_amc(wb, result, meta)

    elif job_type == "project_outdoor":
        _build_ce_project_outdoor(wb, result, meta)
        _build_boq_sheet(wb, result, meta)

    elif job_type == "project_indoor":
        _build_ce_project_indoor(wb, result, meta)
        _build_boq_sheet(wb, result, meta)

    # Write to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
